# Description: This software is designed to
# automate managing the RMA Data entry portion of the Ascom RMA Process.
# Written by: Tory Davenport
# Date: 3/21/2018
# Updated: 6/17/2019

from openpyxl import *
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import *
import time
from time import sleep
import sys

#get the total number of arguments passed
total = len(sys.argv)

# Get arguments list
cmdargs = str(sys.argv)

min_value = int(sys.argv[1])
print("You entered:", min_value)
# For debug only --> #print(row1,"\n", row2,"\n", row3,"\n", row4,"\n", row5,"\n", row6,"\n", row7,"\n", row8,"\n", row9,"\n", row10,"\n")

# Open workbook, create sheet object 
try:
	workbook_obj = load_workbook("..\..\..\FirstLight\Michael Wells - Wegmans\Inventory\Daily Wegmans Inventory Phones.xlsx")
	sheet_obj = workbook_obj["RMA_Phone_Tracking"]
except PermissionError:
	print("Please close the file while running this script.")
except NameError:
	print("")
	
# Create a class to handle row data
class Data_Row:
		
	# Initialize row object
	def __init__(self, row_num):
		row_num = row_num
		# Use to pull individual cells for a given row
		self.model = sheet_obj.cell(row = row_num, column = 4)
		self.serial = sheet_obj.cell(row = row_num, column = 2)
		self.problem_one = sheet_obj.cell(row = row_num, column = 5)
		self.problem_two = sheet_obj.cell(row = row_num, column = 6)
		self.case_number = sheet_obj.cell(row = row_num, column = 8)
		self.notes = sheet_obj.cell(row = row_num, column = 16)
		# Use to print entire rows with formatting	
		self.row_formatted = f"{self.model.value} {self.serial.value} {self.problem_one.value} {self.problem_two.value} {self.case_number.value} {self.notes.value}"

row1 = Data_Row(int(min_value))
row2 = Data_Row(int(min_value+1))
row3 = Data_Row(int(min_value+2))
row4 = Data_Row(int(min_value+3))
row5 = Data_Row(int(min_value+4))
row6 = Data_Row(int(min_value+5))
row7 = Data_Row(int(min_value+6))
row8 = Data_Row(int(min_value+7))
row9 = Data_Row(int(min_value+8))
row10 = Data_Row(int(min_value+9))

workbook_obj.close()

# Launch chrome's web driver 
driver = webdriver.Chrome()
# Bring up ascom form
driver.get("http://www.ascomwireless.com/rma")

assert "Return Material Authorization Form" in driver.title
#####################################################################################
# Fill form information                          									#
#####################################################################################
first_name = driver.find_element_by_name("txtReqFirstName")
first_name.send_keys("Tory")

last_name = driver.find_element_by_name("txtReqLastName")
last_name.send_keys("Davenport")

company = driver.find_element_by_name("txtReqCompany")
company.send_keys("First Light")

customer_number = driver.find_element_by_name("txtAscomCustomerNumber")
customer_number.send_keys("AC3219")

end_user = driver.find_element_by_name("txtReqEndUserName")
end_user.send_keys("Wegmans")

phone_num = driver.find_element_by_name("txtReqPhone")
phone_num.send_keys("5854336649")

email = driver.find_element_by_name("txtReqEmail")
email.send_keys("wegmanstickets@firstlight.net")

protection_plan = driver.find_element_by_name("txtReqPPP")
protection_plan.send_keys("PN-002484")

ship_company_name = driver.find_element_by_name("txtShipCompanyName")
ship_company_name.send_keys("First Light")

ship_street = driver.find_element_by_name("txtShipStreet1")
ship_street.send_keys("7890 Lehigh Crossing")

ship_city = driver.find_element_by_name("txtShipCity")
ship_city.send_keys("Victor")

ship_state = driver.find_element_by_name("txtShipState")
ship_state.send_keys("NY")

ship_zip = driver.find_element_by_name("txtShipZip")
ship_zip.send_keys("14564")

#####################################################################################
# Fill in Form DATA (This will eventually be automated by pulling data from excel)  #
#####################################################################################


# MODEL TYPE 1 - 10
row_1_model_1 = driver.find_element_by_name("ddlModelType1")
row_1_model_1.send_keys(f"{row1.model.value}")

row_2_model_2 = driver.find_element_by_name("ddlModelType2")
row_2_model_2.send_keys(f"{row2.model.value}")

row_3_model_3 = driver.find_element_by_name("ddlModelType3")
row_3_model_3.send_keys(f"{row3.model.value}")

row_4_model_4 = driver.find_element_by_name("ddlModelType4")
row_4_model_4.send_keys(f"{row4.model.value}")

row_5_model_5 = driver.find_element_by_name("ddlModelType5")
row_5_model_5.send_keys(f"{row5.model.value}")

row_6_model_6 = driver.find_element_by_name("ddlModelType6")
row_6_model_6.send_keys(f"{row6.model.value}")

row_7_model_7 = driver.find_element_by_name("ddlModelType7")
row_7_model_7.send_keys(f"{row7.model.value}")

row_8_model_8 = driver.find_element_by_name("ddlModelType8")
row_8_model_8.send_keys(f"{row8.model.value}")

row_9_model_9 = driver.find_element_by_name("ddlModelType9")
row_9_model_9.send_keys(f"{row9.model.value}")

row_10_model_10 = driver.find_element_by_name("ddlModelType10")
row_10_model_10.send_keys(f"{row10.model.value}")

# Fill Serial's
serial_1 = driver.find_element_by_name("txtSerial1")
serial_1.send_keys(f"{row1.serial.value}")

serial_2 = driver.find_element_by_name("txtSerial2")
serial_2.send_keys(f"{row2.serial.value}")

serial_3 = driver.find_element_by_name("txtSerial3")
serial_3.send_keys(f"{row3.serial.value}")

serial_4 = driver.find_element_by_name("txtSerial4")
serial_4.send_keys(f"{row4.serial.value}")

serial_5 = driver.find_element_by_name("txtSerial5")
serial_5.send_keys(f"{row5.serial.value}")

serial_6 = driver.find_element_by_name("txtSerial6")
serial_6.send_keys(f"{row6.serial.value}")

serial_7 = driver.find_element_by_name("txtSerial7")
serial_7.send_keys(f"{row7.serial.value}")

serial_8 = driver.find_element_by_name("txtSerial8")
serial_8.send_keys(f"{row8.serial.value}")

serial_9 = driver.find_element_by_name("txtSerial9")
serial_9.send_keys(f"{row9.serial.value}")

serial_10 = driver.find_element_by_name("txtSerial10")
serial_10.send_keys(f"{row10.serial.value}")


# Required Variables
# This allows you to enter two problem descriptions per item

# If statement for error checking, if the cell is blank it will print a None type value, this will remove the none value when inputting data into form.
if row1.problem_two.value == None:
	row1.problem_two.value = " "
if row2.problem_two.value == None:
	row2.problem_two.value = " "
if row3.problem_two.value == None:
	row3.problem_two.value = " "
if row4.problem_two.value == None:
	row4.problem_two.value = " "
if row5.problem_two.value == None:
	row5.problem_two.value = " "
if row6.problem_two.value == None:
	row6.problem_two.value = " "
if row7.problem_two.value == None:
	row7.problem_two.value = " "
if row8.problem_two.value == None:
	row8.problem_two.value = " "
if row9.problem_two.value == None:
	row9.problem_two.value = " "
if row10.problem_two.value == None:
	row10.problem_two.value = " "

# Populate Problem 1 and problem 2 for each row in form
problem_1 = [f'{row1.problem_one.value}',f'{row1.problem_two.value}']
problem_2 = [f'{row2.problem_one.value}',f'{row2.problem_two.value}']
problem_3 = [f'{row3.problem_one.value}',f'{row3.problem_two.value}']
problem_4 = [f'{row4.problem_one.value}',f'{row4.problem_two.value}']
problem_5 = [f'{row5.problem_one.value}',f'{row5.problem_two.value}']
problem_6 = [f'{row6.problem_one.value}',f'{row6.problem_two.value}']
problem_7 = [f'{row7.problem_one.value}',f'{row7.problem_two.value}']
problem_8 = [f'{row8.problem_one.value}',f'{row8.problem_two.value}']
problem_9 = [f'{row9.problem_one.value}',f'{row9.problem_two.value}']
problem_10 = [f'{row10.problem_one.value}',f'{row10.problem_two.value}']

# Fill Problem Descriptions 1 and 2 for each of the 10 rows
problem_descr1_row1 = driver.find_element_by_name("ddlProblemDescr1_row1")
problem_descr1_row1.send_keys(problem_1[0])

problem_descr2_row1 = driver.find_element_by_name("ddlProblemDescr2_row1")
problem_descr2_row1.send_keys(problem_1[1])

problem_descr1_row2 = driver.find_element_by_name("ddlProblemDescr1_row2")
problem_descr1_row2.send_keys(problem_2[0])

problem_descr2_row2 = driver.find_element_by_name("ddlProblemDescr2_row2")
problem_descr2_row2.send_keys(problem_2[1])

problem_descr1_row3 = driver.find_element_by_name("ddlProblemDescr1_row3")
problem_descr1_row3.send_keys(problem_3[0])

problem_descr2_row3 = driver.find_element_by_name("ddlProblemDescr2_row3")
problem_descr2_row3.send_keys(problem_3[1])

problem_descr1_row4 = driver.find_element_by_name("ddlProblemDescr1_row4")
problem_descr1_row4.send_keys(problem_4[0])

problem_descr2_row4 = driver.find_element_by_name("ddlProblemDescr2_row4")
problem_descr2_row4.send_keys(problem_4[1])

problem_descr1_row5 = driver.find_element_by_name("ddlProblemDescr1_row5")
problem_descr1_row5.send_keys(problem_5[0])

problem_descr2_row5 = driver.find_element_by_name("ddlProblemDescr2_row5")
problem_descr2_row5.send_keys(problem_5[1])

problem_descr1_row6 = driver.find_element_by_name("ddlProblemDescr1_row6")
problem_descr1_row6.send_keys(problem_6[0])

problem_descr2_row6 = driver.find_element_by_name("ddlProblemDescr2_row6")
problem_descr2_row6.send_keys(problem_6[1])

problem_descr1_row7 = driver.find_element_by_name("ddlProblemDescr1_row7")
problem_descr1_row7.send_keys(problem_7[0])

problem_descr2_row7 = driver.find_element_by_name("ddlProblemDescr2_row7")
problem_descr2_row7.send_keys(problem_7[1])

problem_descr1_row8 = driver.find_element_by_name("ddlProblemDescr1_row8")
problem_descr1_row8.send_keys(problem_8[0])

problem_descr2_row8 = driver.find_element_by_name("ddlProblemDescr2_row8")
problem_descr2_row8.send_keys(problem_8[1])

problem_descr1_row9 = driver.find_element_by_name("ddlProblemDescr1_row9")
problem_descr1_row9.send_keys(problem_9[0])

problem_descr2_row9 = driver.find_element_by_name("ddlProblemDescr2_row9")
problem_descr2_row9.send_keys(problem_9[1])

problem_descr1_row10 = driver.find_element_by_name("ddlProblemDescr1_row10")
problem_descr1_row10.send_keys(problem_10[0])

problem_descr2_row10 = driver.find_element_by_name("ddlProblemDescr2_row10")
problem_descr2_row10.send_keys(problem_10[1])

# Fill Case number's
case_1 = driver.find_element_by_name("txtCase_row1")
if row1.problem_one.value == "One way audio" or row1.problem_two.value == "One way audio":
	case_1.send_keys("CUS-20104")
else:
	case_1.send_keys(" ")

case_2 = driver.find_element_by_name("txtCase_row2")
if row2.problem_one.value == "One way audio" or row2.problem_two.value == "One way audio":
	case_2.send_keys("CUS-20104")
else:
	case_2.send_keys(" ")

case_3 = driver.find_element_by_name("txtCase_row3")
if row3.problem_one.value == "One way audio" or row3.problem_two.value == "One way audio":
	case_3.send_keys("CUS-20104")
else:
	case_3.send_keys(" ")

case_4 = driver.find_element_by_name("txtCase_row4")
if row4.problem_one.value == "One way audio" or row4.problem_two.value == "One way audio":
	case_4.send_keys("CUS-20104")
else:
	case_4.send_keys(" ")

case_5 = driver.find_element_by_name("txtCase_row5")
if row5.problem_one.value == "One way audio" or row5.problem_two.value == "One way audio":
	case_5.send_keys("CUS-20104")
else:
	case_5.send_keys(" ")

case_6 = driver.find_element_by_name("txtCase_row6")
if row6.problem_one.value == "One way audio" or row6.problem_two.value == "One way audio":
	case_6.send_keys("CUS-20104")
else:
	case_6.send_keys(" ")

case_7 = driver.find_element_by_name("txtCase_row7")
if row7.problem_one.value == "One way audio" or row7.problem_two.value == "One way audio":
	case_7.send_keys("CUS-20104")
else:
	case_7.send_keys(" ")

case_8 = driver.find_element_by_name("txtCase_row8")
if row8.problem_one.value == "One way audio" or row8.problem_two.value == "One way audio":
	case_8.send_keys("CUS-20104")
else:
	case_8.send_keys(" ")

case_9 = driver.find_element_by_name("txtCase_row9")
if row9.problem_one.value == "One way audio" or row9.problem_two.value == "One way audio":
	case_9.send_keys("CUS-20104")
else:
	case_9.send_keys(" ")

case_10 = driver.find_element_by_name("txtCase_row10")
if row10.problem_one.value == "One way audio" or row10.problem_two.value == "One way audio":
	case_10.send_keys("CUS-20104")
else:
	case_10.send_keys(" ")

# Import and fill notes column into notes section of form
notes = driver.find_element_by_name("txtNotesComments")
notes.send_keys(f"Serial: {row1.serial.value}, Notes: {row1.notes.value}\n"f"Serial: {row2.serial.value}, Notes: {row2.notes.value}\n"f"Serial: {row3.serial.value}, Notes: {row3.notes.value}\n"f"Serial: {row4.serial.value}, Notes: {row4.notes.value}\n"f"Serial: {row5.serial.value}, Notes: {row5.notes.value}\n"f"Serial: {row6.serial.value}, Notes: {row6.notes.value}\n"f"Serial: {row7.serial.value}, Notes: {row7.notes.value}\n"f"Serial: {row8.serial.value}, Notes: {row8.notes.value}\n"f"Serial: {row9.serial.value}, Notes: {row9.notes.value}\n"f"Serial: {row10.serial.value}, Notes: {row10.notes.value}\n")

# Actions to print form (disable to for testing to avoid submissions)
#driver.find_element_by_name("btnPrint").click()

input("press enter")
driver.quit()


